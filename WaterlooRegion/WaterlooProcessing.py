from openpyxl import Workbook

wb = Workbook()

ws_a = wb.active
ws_a.title = 'sheet_1'
ws_b = wb.create_sheet('sheet_2')
ws_c = wb.create_sheet('sheet_3')
ws_d = wb.create_sheet('sheet_4')

for item in wb:
    print(item.title)

for counter in range(1,101):
    for x in range(1,101):
        for y in range(1,101):
            ws_a.cell(row=x,column=y,value=counter)

col_c = ws_a['C']
for tow in ws_a.



wb.save('template_1.xlsx')